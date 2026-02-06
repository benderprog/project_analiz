import logging
from dataclasses import dataclass
from typing import Iterable

from django.db import transaction
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class ClassifierRow:
    event_type: str
    pattern: str
    article_of_law: str


@dataclass
class ImportSummary:
    created_types: int = 0
    created_patterns: int = 0
    updated_patterns: int = 0
    skipped_rows: int = 0


def _normalize_cell(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_classifier_xlsx(file_obj) -> tuple[list[ClassifierRow], int]:
    """Parse classifier XLSX with strict 3-column format and optional type fill-down."""
    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    rows: list[ClassifierRow] = []
    skipped_rows = 0
    last_event_type = ""

    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_event_type = _normalize_cell(row[0] if len(row) > 0 else None)
        raw_pattern = _normalize_cell(row[1] if len(row) > 1 else None)
        raw_article = _normalize_cell(row[2] if len(row) > 2 else None)

        if not raw_event_type and not raw_pattern and not raw_article:
            skipped_rows += 1
            continue

        event_type = raw_event_type or (last_event_type if (raw_pattern or raw_article) else "")
        if not event_type:
            skipped_rows += 1
            continue

        rows.append(
            ClassifierRow(
                event_type=event_type,
                pattern=raw_pattern,
                article_of_law=raw_article,
            )
        )
        last_event_type = event_type

    logger.info("Parsed %s classifier rows (skipped %s)", len(rows), skipped_rows)
    return rows, skipped_rows


def import_classifier_rows(
    rows: Iterable[ClassifierRow],
    *,
    clear_before: bool = False,
    skipped_rows: int = 0,
) -> ImportSummary:
    """Import classifier rows into the DB."""
    from apps.classifier.models import EventType, EventTypePattern

    summary = ImportSummary(skipped_rows=skipped_rows)

    with transaction.atomic():
        if clear_before:
            EventType.objects.all().delete()

        for row in rows:
            event_type_value = row.event_type.strip()
            event_type_obj, created = EventType.objects.get_or_create(
                event_type=event_type_value
            )
            if created:
                summary.created_types += 1

            pattern_value = row.pattern.strip()
            if not pattern_value:
                continue

            pattern_obj, created = EventTypePattern.objects.get_or_create(
                event_type=event_type_obj,
                pattern=pattern_value,
                defaults={"article_of_law": row.article_of_law.strip() or None},
            )
            if created:
                summary.created_patterns += 1
                continue

            if row.article_of_law.strip() and row.article_of_law.strip() != (
                pattern_obj.article_of_law or ""
            ):
                pattern_obj.article_of_law = row.article_of_law.strip()
                pattern_obj.save(update_fields=["article_of_law"])
                summary.updated_patterns += 1

    logger.info(
        "Imported classifier rows: types=%s patterns=%s updated=%s skipped=%s",
        summary.created_types,
        summary.created_patterns,
        summary.updated_patterns,
        summary.skipped_rows,
    )
    return summary
