from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook
from docx import Document

from apps.portaldb.models import Event, Offender, Pu, Subdivision

PORTAL_DB_ALIAS = "portal"
DEFAULT_XLSX = "subdivizion_primer.xlsx"
DEFAULT_DOCX = "test_svodka_semantic.docx"
DEFAULT_DOB = date(1900, 1, 1)


@dataclass(frozen=True)
class OffenderData:
    first_name: str
    second_name: str
    patronymic_name: str
    date_of_birth: date


@dataclass(frozen=True)
class Scenario:
    match_type: str
    time_shift_minutes: int
    offenders_mode: str
    event_type: str
    article_of_law: str


SCENARIOS: dict[int, Scenario] = {
    1: Scenario(
        match_type="FULL",
        time_shift_minutes=0,
        offenders_mode="all",
        event_type="Проверка режима",
        article_of_law="—",
    ),
    3: Scenario(
        match_type="PARTIAL",
        time_shift_minutes=40,
        offenders_mode="first_only",
        event_type="",
        article_of_law="",
    ),
    7: Scenario(
        match_type="PARTIAL",
        time_shift_minutes=0,
        offenders_mode="none",
        event_type="",
        article_of_law="",
    ),
    10: Scenario(
        match_type="FULL",
        time_shift_minutes=0,
        offenders_mode="all",
        event_type="Проверка режима",
        article_of_law="—",
    ),
    14: Scenario(
        match_type="PARTIAL",
        time_shift_minutes=15,
        offenders_mode="none",
        event_type="Несоответствующий тип",
        article_of_law="",
    ),
}

NORMALIZATION_REPLACEMENTS = {
    "ПОГЗ-2": "ПОГЗ №2",
    "ПОГО-Южное": "ПОГО «Южное»",
    "ОП-Центральное": "ОП «Центральное»",
    "ПЗ1": "ПЗ №1",
}


class Command(BaseCommand):
    help = "Seed portal_db with PU/subdivisions and curated events for upload/analyze testing."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--xlsx",
            default=str(Path(settings.BASE_DIR) / DEFAULT_XLSX),
            help="Path to subdivizion_primer.xlsx (defaults to project root).",
        )
        parser.add_argument(
            "--docx",
            default=str(Path(settings.BASE_DIR) / DEFAULT_DOCX),
            help="Path to test_svodka_semantic.docx (defaults to project root).",
        )
        parser.add_argument(
            "--reset",
            action="store_true",
            help="Wipe portal tables (offenders/events/subdivisions/pu) before seeding.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Print planned operations without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        xlsx_path = Path(options["xlsx"]).expanduser()
        docx_path = Path(options["docx"]).expanduser()
        dry_run = options["dry_run"]

        if not xlsx_path.exists():
            raise CommandError(f"XLSX file not found: {xlsx_path}")
        if not docx_path.exists():
            raise CommandError(f"DOCX file not found: {docx_path}")

        if options["reset"]:
            self._reset_tables(dry_run=dry_run)

        pu_rows = self._read_xlsx(xlsx_path)
        pu_stats, subdivision_stats, subdivision_candidates = self._seed_pu_subdivisions(
            pu_rows, dry_run=dry_run
        )

        paragraphs = self._read_docx(docx_path)
        event_stats = self._seed_events(
            paragraphs,
            subdivision_candidates,
            dry_run=dry_run,
        )

        self._print_summary(
            pu_stats=pu_stats,
            subdivision_stats=subdivision_stats,
            event_stats=event_stats,
            dry_run=dry_run,
        )

    def _reset_tables(self, *, dry_run: bool) -> None:
        if dry_run:
            self.stdout.write("[DRY RUN] Would delete offenders, events, subdivisions, and PUs.")
            return

        Offender.objects.using(PORTAL_DB_ALIAS).all().delete()
        Event.objects.using(PORTAL_DB_ALIAS).all().delete()
        Subdivision.objects.using(PORTAL_DB_ALIAS).all().delete()
        Pu.objects.using(PORTAL_DB_ALIAS).all().delete()
        self.stdout.write("Portal tables reset.")

    def _read_xlsx(self, path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise CommandError("XLSX file is empty.")

        headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
        required = {"short_name", "full_name", "pu"}
        missing = required - set(headers)
        if missing:
            raise CommandError(f"XLSX file missing required columns: {', '.join(sorted(missing))}")

        header_index = {header: idx for idx, header in enumerate(headers)}
        data_rows: list[dict[str, str]] = []
        for row in rows[1:]:
            short_name = self._cell_value(row, header_index["short_name"])
            full_name = self._cell_value(row, header_index["full_name"])
            pu_name = self._cell_value(row, header_index["pu"])
            if not (short_name or full_name or pu_name):
                continue
            data_rows.append(
                {
                    "short_name": short_name,
                    "full_name": full_name,
                    "pu": pu_name,
                }
            )
        return data_rows

    def _cell_value(self, row: tuple, index: int) -> str:
        try:
            value = row[index]
        except IndexError:
            return ""
        if value is None:
            return ""
        return str(value).strip()

    def _seed_pu_subdivisions(
        self, rows: list[dict[str, str]], *, dry_run: bool
    ) -> tuple[dict[str, int], dict[str, int], list[dict[str, str | Subdivision]]]:
        pu_created = 0
        pu_existing = 0
        subdivision_created = 0
        subdivision_existing = 0

        pu_cache: dict[str, Pu] = {}
        subdivision_candidates: list[dict[str, str | Subdivision]] = []
        planned_subdivision_keys: set[tuple[str, str]] = set()

        for row in rows:
            pu_name = row["pu"].strip()
            if not pu_name:
                continue

            pu = pu_cache.get(pu_name)
            if pu is None:
                pu_qs = Pu.objects.using(PORTAL_DB_ALIAS).filter(short_name=pu_name)
                pu = pu_qs.first()
                if pu:
                    pu_existing += 1
                else:
                    pu_created += 1
                    if not dry_run:
                        pu = Pu.objects.using(PORTAL_DB_ALIAS).create(
                            short_name=pu_name, full_name=pu_name
                        )
                    else:
                        pu = Pu(short_name=pu_name, full_name=pu_name)
                pu_cache[pu_name] = pu

            subdivision_name = row["full_name"].strip() or row["short_name"].strip()
            if not subdivision_name:
                continue
            subdivision_key = (subdivision_name, pu_name)
            if subdivision_key in planned_subdivision_keys:
                continue

            subdivision = Subdivision.objects.using(PORTAL_DB_ALIAS).filter(
                name=subdivision_name, parent_pu=pu
            ).first()
            if subdivision:
                subdivision_existing += 1
            else:
                subdivision_created += 1
                if not dry_run:
                    subdivision = Subdivision.objects.using(PORTAL_DB_ALIAS).create(
                        name=subdivision_name, parent_pu=pu
                    )
                else:
                    subdivision = Subdivision(name=subdivision_name, parent_pu=pu)

            planned_subdivision_keys.add(subdivision_key)
            subdivision_candidates.append(
                {
                    "name": subdivision_name,
                    "short_name": row["short_name"].strip(),
                    "subdivision": subdivision,
                }
            )

        return (
            {"created": pu_created, "existing": pu_existing},
            {"created": subdivision_created, "existing": subdivision_existing},
            subdivision_candidates,
        )

    def _read_docx(self, path: Path) -> list[str]:
        document = Document(path)
        paragraphs = []
        for paragraph in document.paragraphs:
            text = paragraph.text.strip()
            if text:
                paragraphs.append(text)
        return paragraphs

    def _seed_events(
        self,
        paragraphs: list[str],
        subdivision_candidates: list[dict[str, str | Subdivision]],
        *,
        dry_run: bool,
    ) -> dict[str, object]:
        inserted_events: list[int] = []
        offender_count = 0
        skipped_missing_subdivision: list[int] = []

        for index, text in enumerate(paragraphs, start=1):
            scenario = SCENARIOS.get(index)
            if not scenario:
                continue

            event_datetime = self._extract_datetime(text)
            if not event_datetime:
                continue
            event_datetime += timedelta(minutes=scenario.time_shift_minutes)

            subdivision = self._match_subdivision(text, subdivision_candidates)
            if subdivision is None:
                skipped_missing_subdivision.append(index)
                continue

            offenders = self._extract_offenders(text)
            offenders = self._apply_offender_mode(offenders, scenario.offenders_mode)

            if not dry_run:
                event = Event.objects.using(PORTAL_DB_ALIAS).create(
                    date_detection=event_datetime,
                    find_subdivision_unit=subdivision,
                    event_type=scenario.event_type,
                    article_of_law=scenario.article_of_law,
                )
                for offender in offenders:
                    Offender.objects.using(PORTAL_DB_ALIAS).create(
                        event=event,
                        first_name=offender.first_name,
                        second_name=offender.second_name,
                        patronymic_name=offender.patronymic_name,
                        date_of_birth=offender.date_of_birth,
                    )

            inserted_events.append(index)
            offender_count += len(offenders)

        return {
            "inserted_events": inserted_events,
            "offenders_inserted": offender_count,
            "skipped_missing_subdivision": skipped_missing_subdivision,
        }

    def _extract_datetime(self, text: str) -> datetime | None:
        date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
        time_match = re.search(r"(\d{1,2})[.:](\d{2})", text)
        if not date_match or not time_match:
            return None
        date_part = datetime.strptime(date_match.group(1), "%d.%m.%Y").date()
        hour = int(time_match.group(1))
        minute = int(time_match.group(2))
        return datetime.combine(date_part, datetime.min.time()).replace(hour=hour, minute=minute)

    def _normalize_text(self, text: str) -> str:
        normalized = text
        for source, target in NORMALIZATION_REPLACEMENTS.items():
            normalized = normalized.replace(source, target)
        return normalized

    def _match_subdivision(
        self, text: str, subdivision_candidates: list[dict[str, str | Subdivision]]
    ) -> Subdivision | None:
        normalized = self._normalize_text(text).lower()
        best_match = None
        best_length = -1
        for candidate in subdivision_candidates:
            name = str(candidate["name"]).lower()
            short_name = str(candidate["short_name"]).lower()
            for token in {name, short_name}:
                if token and token in normalized:
                    if len(token) > best_length:
                        best_match = candidate["subdivision"]
                        best_length = len(token)
        if isinstance(best_match, Subdivision):
            return best_match
        return None

    def _extract_offenders(self, text: str) -> list[OffenderData]:
        offenders: list[OffenderData] = []
        name_pattern = re.compile(
            r"\b[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+)?",
            re.UNICODE,
        )
        for match in name_pattern.finditer(text):
            full_name = match.group(0)
            tokens = full_name.split()
            if len(tokens) < 2:
                continue
            second_name = tokens[0]
            first_name = tokens[1]
            patronymic_name = tokens[2] if len(tokens) > 2 else ""
            dob = self._extract_birth_date(text, match.end())
            offenders.append(
                OffenderData(
                    first_name=first_name,
                    second_name=second_name,
                    patronymic_name=patronymic_name,
                    date_of_birth=dob,
                )
            )
        return offenders

    def _extract_birth_date(self, text: str, start_idx: int) -> date:
        snippet = text[start_idx : start_idx + 40]
        date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", snippet)
        if date_match:
            return datetime.strptime(date_match.group(1), "%d.%m.%Y").date()
        year_match = re.search(r"(19\d{2}|20\d{2})", snippet)
        if year_match:
            return date(int(year_match.group(1)), 1, 1)
        return DEFAULT_DOB

    def _apply_offender_mode(
        self, offenders: list[OffenderData], mode: str
    ) -> list[OffenderData]:
        if mode == "none":
            return []
        if mode == "first_only":
            return offenders[:1]
        return offenders

    def _print_summary(
        self,
        *,
        pu_stats: dict[str, int],
        subdivision_stats: dict[str, int],
        event_stats: dict[str, object],
        dry_run: bool,
    ) -> None:
        inserted_events = event_stats["inserted_events"]
        offenders_inserted = event_stats["offenders_inserted"]
        skipped_missing_subdivision = event_stats["skipped_missing_subdivision"]

        if dry_run:
            self.stdout.write("[DRY RUN] No data was written to the database.")

        self.stdout.write(
            "PU summary: created={created}, existing={existing}".format(**pu_stats)
        )
        self.stdout.write(
            "Subdivision summary: created={created}, existing={existing}".format(
                **subdivision_stats
            )
        )
        self.stdout.write(
            f"Events inserted: {len(inserted_events)} (paragraphs: {inserted_events})"
        )
        if skipped_missing_subdivision:
            self.stdout.write(
                "Skipped events due to missing subdivision: "
                f"{skipped_missing_subdivision}"
            )
        self.stdout.write(f"Offenders inserted: {offenders_inserted}")

        scenario_lines = []
        for index, scenario in SCENARIOS.items():
            scenario_lines.append(f"Paragraph {index}: {scenario.match_type}")
        self.stdout.write("Intended scenarios: " + "; ".join(scenario_lines))
